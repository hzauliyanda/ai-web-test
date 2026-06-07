#!/usr/bin/env python3
"""
Generate XLSX import file from test cases markdown file.

Reads the testcase-template.xlsx template for styling reference, then populates
it with test case data parsed from a markdown file.

The output format matches the template columns:
  模块一, 模块二, 模块三, 模块四, 模块五, 模块六,
  用例名称, 优先级, 负责人, 前置条件, 用例步骤, 预期结果, 关联自动化用例, 备注

Priority in XLSX is always set to 3 (default value for platform import).
Original priority (P0/P1/P2) is preserved in Markdown and XMind formats.
Assignee defaults to <YOUR_ASSIGNEE>.

Usage:
    python generate_xlsx.py test-cases.md
    python generate_xlsx.py test-cases.md output.xlsx
    python generate_xlsx.py test-cases.md --assignee <YOUR_ASSIGNEE>
"""

import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PRIORITY_MAP = {'P0': 0, 'P1': 1, 'P2': 2, 'P3': 3}
XLSX_PRIORITY = 3


def _get_default_assignee():
    assignee = os.environ.get('TESTCASE_DEFAULT_ASSIGNEE', '')
    if assignee:
        return assignee
    config_path = Path.home() / '.config' / 'testcase-generator' / 'config'
    if config_path.exists():
        for line in config_path.read_text(encoding='utf-8').splitlines():
            line = line.strip()
            if line.startswith('default_assignee='):
                return line.split('=', 1)[1].strip()
    return ''


DEFAULT_ASSIGNEE = _get_default_assignee()


def parse_test_cases(md_content):
    lines = md_content.split('\n')
    test_cases = []

    current_modules = []
    current_priority = None
    current_tc = None
    current_section = None

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        if stripped.startswith('## '):
            if current_tc:
                test_cases.append(current_tc)
            module_name = stripped[3:].strip()
            if module_name.endswith('模块'):
                module_name = module_name[:-2]
            current_modules = [module_name]
            current_priority = None
            current_tc = None
            current_section = None
            continue

        if stripped.startswith('### P') and current_modules:
            if current_tc:
                test_cases.append(current_tc)
            priority_str = stripped.replace('### ', '').split()[0].strip()
            current_priority = PRIORITY_MAP.get(priority_str)
            current_tc = None
            current_section = None
            continue

        if stripped.startswith('#### TC-') and current_modules and current_priority is not None:
            if current_tc:
                test_cases.append(current_tc)
            match = re.match(r'#### (TC-[\w-]+):\s*(.+)', stripped)
            if match:
                tc_id = match.group(1)
                tc_title = match.group(2).strip()
                current_tc = {
                    'modules': list(current_modules),
                    'name': f'{tc_id}: {tc_title}',
                    'priority': current_priority,
                    'prerequisites': [],
                    'steps': [],
                    'expected_results': [],
                    'implementation_ref': '',
                    'description': '',
                    'assignee': DEFAULT_ASSIGNEE,
                    '_section': None,
                }
            else:
                current_tc = None
            current_section = None
            continue

        if current_tc is None:
            continue

        if '**优先级' in stripped or '**Priority' in stripped:
            continue
        if '**测试类型' in stripped or '**Test Type' in stripped:
            continue

        if '**描述' in stripped or '**Description' in stripped:
            desc = re.sub(r'\*\*描述\*\*[：:]\s*', '', stripped)
            desc = re.sub(r'\*\*Description\*\*[：:]\s*', '', desc)
            desc = desc.replace('**', '').strip()
            if desc:
                current_tc['description'] = desc
            current_section = None
            continue

        if '**前置条件' in stripped or '**Prerequisites' in stripped:
            current_section = 'prerequisites'
            continue
        if '**测试步骤' in stripped or '**Test Steps' in stripped:
            current_section = 'steps'
            continue
        if '**预期结果' in stripped or '**Expected Result' in stripped:
            current_section = 'expected_results'
            continue

        if '**实现参考' in stripped or '**Implementation Reference' in stripped:
            current_section = 'impl'
            impl = re.sub(r'\*\*实现参考\*\*[：:]\s*', '', stripped)
            impl = re.sub(r'\*\*Implementation Reference\*\*[：:]\s*', '', impl)
            impl = impl.replace('**', '').strip()
            if impl:
                current_tc['implementation_ref'] = impl
            continue

        is_list_item = stripped.startswith('- ') or re.match(r'^\d+\.\s', stripped)
        if is_list_item and current_section:
            item_text = re.sub(r'^\d+\.\s', '', stripped)
            item_text = item_text.lstrip('- ').strip().strip('-').strip()

            if not item_text:
                continue

            if current_section == 'prerequisites':
                current_tc['prerequisites'].append(item_text)
            elif current_section == 'steps':
                current_tc['steps'].append(item_text)
            elif current_section == 'expected_results':
                current_tc['expected_results'].append(item_text)
            elif current_section == 'impl':
                current_tc['implementation_ref'] += '\n' + item_text if current_tc['implementation_ref'] else item_text
            continue

        if current_section == 'expected_results' and stripped.startswith('|') and '|' in stripped[1:]:
            cells = [c.strip() for c in stripped.split('|') if c.strip()]
            skip_words = {'✅', '验证', 'Cin7 字段', 'Shopline 字段', 'Cin7字段', 'Shopline字段'}
            for cell in cells:
                if cell and cell not in skip_words:
                    current_tc['expected_results'].append(cell)
            continue

    if current_tc:
        test_cases.append(current_tc)

    return test_cases


def generate_xlsx(test_cases, output_path, template_path=None, assignee=DEFAULT_ASSIGNEE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '测试用例'

    headers_module = ['模块一', '模块二', '模块三', '模块四', '模块五', '模块六']
    headers_data = ['用例名称', '优先级', '负责人', '前置条件', '用例步骤', '预期结果', '关联自动化用例', '备注']
    all_headers = headers_module + headers_data

    col_widths = {
        1: 14, 2: 14, 3: 14, 4: 14, 5: 14, 6: 14,
        7: 22, 8: 8, 9: 14, 10: 25, 11: 28, 12: 58, 13: 24, 14: 13,
    }

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if col_idx <= MODULE_COLUMNS:
            cell.font = Font(name='宋体', bold=True, size=11, charset=134)
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        else:
            cell.font = Font(name='方正书宋_GBK', bold=True, size=10, charset=134)
            cell.fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row_idx = 2
    for tc in test_cases:
        modules = list(tc['modules'])
        while len(modules) < MODULE_COLUMNS:
            modules.append('')

        for col_idx in range(MODULE_COLUMNS):
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=modules[col_idx])
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT

        cell = ws.cell(row=row_idx, column=7, value=tc['name'])
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT

        cell = ws.cell(row=row_idx, column=8, value=XLSX_PRIORITY)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell = ws.cell(row=row_idx, column=9, value=assignee)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT

        prerequisites_text = '\n'.join(tc['prerequisites'])
        cell = ws.cell(row=row_idx, column=10, value=prerequisites_text)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT

        steps_text = '\n'.join(
            f'{i+1}. {step}' if not step.startswith('继') else step
            for i, step in enumerate(tc['steps'])
        )
        cell = ws.cell(row=row_idx, column=11, value=steps_text)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT

        expected_items = list(tc['expected_results'])
        if tc.get('implementation_ref'):
            expected_items.append(f'【实现参考】{tc["implementation_ref"]}')
        expected_text = '\n'.join(expected_items)
        cell = ws.cell(row=row_idx, column=12, value=expected_text)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT

        cell = ws.cell(row=row_idx, column=13, value='')
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT

        desc_text = tc.get('description', '')
        cell = ws.cell(row=row_idx, column=14, value=desc_text)
        cell.font = DATA_FONT
        cell.alignment = DATA_ALIGNMENT

        row_idx += 1

    ws.freeze_panes = 'A2'

    wb.save(str(output_path))
    return output_path


DATA_FONT = Font(name='宋体', size=10, charset=134)
DATA_ALIGNMENT = Alignment(horizontal='left', vertical='center', wrapText=True)
MODULE_COLUMNS = 6


def count_test_cases(md_content):
    lines = md_content.split('\n')
    total = 0
    p_counts = {'P0': 0, 'P1': 0, 'P2': 0, 'P3': 0}
    current_priority = None

    for line in lines:
        stripped = line.strip()
        if stripped.startswith('### P'):
            priority_str = stripped.replace('### ', '').split()[0].strip()
            current_priority = priority_str
        elif stripped.startswith('#### TC-') and current_priority:
            total += 1
            if current_priority in p_counts:
                p_counts[current_priority] += 1

    return total, p_counts


def main():
    if len(sys.argv) < 2:
        print("=" * 60)
        print("XLSX Test Case Generator v1.0")
        print("=" * 60)
        print("\nUsage:")
        print("  python generate_xlsx.py <markdown_file> [output_file] [options]")
        print("\nOptions:")
        print("  --assignee NAME    Default assignee (configurable via env or config file)")
        print("\nConfiguration:")
        print("  Set default assignee via:")
        print("    - Environment variable: TESTCASE_DEFAULT_ASSIGNEE")
        print("    - Config file: ~/.config/testcase-generator/config (default_assignee=NAME)")
        print("\nExamples:")
        print("  python generate_xlsx.py test-cases.md")
        print("  python generate_xlsx.py test-cases.md output.xlsx")
        print("  python generate_xlsx.py test-cases.md --assignee your_name")
        print("\nColumn mapping:")
        print("  模块一~模块六: Module hierarchy from markdown H2 headers")
        print("  用例名称: TC-ID + title")
        print("  优先级: Always 3 (default for platform import)")
        print("  负责人: Configurable via --assignee, env var, or config file")
        print("  前置条件: Prerequisites")
        print("  用例步骤: Test steps (numbered)")
        print("  预期结果: Expected results + implementation reference")
        print("  关联自动化用例: Left empty")
        print("  备注: Test description")
        print("=" * 60)
        sys.exit(1)

    md_file = sys.argv[1]
    output_file = None
    assignee = DEFAULT_ASSIGNEE

    i = 2
    while i < len(sys.argv):
        if sys.argv[i] == '--assignee' and i + 1 < len(sys.argv):
            assignee = sys.argv[i + 1]
            i += 2
        elif sys.argv[i].endswith('.xlsx'):
            output_file = sys.argv[i]
            i += 1
        else:
            i += 1

    md_path = Path(md_file)
    if not md_path.exists():
        print(f"❌ Error: File not found: {md_path}")
        sys.exit(1)

    DEFAULT_OUTPUT_DIR = Path.home() / 'AI-TEST' / 'TestCase'
    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if output_file is None:
        output_path = DEFAULT_OUTPUT_DIR / md_path.with_suffix('.xlsx').name
    else:
        output_path = Path(output_file)

    with open(md_path, 'r', encoding='utf-8') as f:
        md_content = f.read()

    print(f"📖 Reading: {md_path}")

    total, p_counts = count_test_cases(md_content)
    test_cases = parse_test_cases(md_content)

    print("\n" + "=" * 60)
    print("XLSX Test Cases Summary")
    print("=" * 60)
    print(f"📝 Total Test Cases: {len(test_cases)} (found {total} in markdown)")
    print(f"\nPriority Distribution:")
    print(f"  🔴 P0 (Critical): {sum(1 for tc in test_cases if tc['priority'] == 0)}")
    print(f"  🟠 P1 (Important): {sum(1 for tc in test_cases if tc['priority'] == 1)}")
    print(f"  🔵 P2 (Normal): {sum(1 for tc in test_cases if tc['priority'] == 2)}")
    print(f"  ⚪ P3 (Low): {sum(1 for tc in test_cases if tc['priority'] == 3)}")
    print("=" * 60)

    generate_xlsx(test_cases, output_path, assignee=assignee)

    import os
    print(f"\n✅ Successfully created XLSX file: {output_path}")
    print(f"   File size: {os.path.getsize(output_path)} bytes")
    print(f"   Assignee: {assignee}")


if __name__ == '__main__':
    main()